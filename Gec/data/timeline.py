# encoding: utf-8

"""
project = zlr数据处理
file_name = Timeline
author = Administrator
datetime = 2020/5/7 0007 下午 15:42
from = office desktop
"""
from openpyxl import load_workbook
from Calf.data import BaseModel
from Gec import project_dir
from Gec.exception import ExceptionInfo


class Timeline:

    def __init__(self, name):
        self.bm = BaseModel(
            tn='qcc',
            location='gcxy',
            dbname='data'
        )
        self.name = name
        self.timeline = []
        self.getTimeline()
        self.timeline.sort(
            key=lambda x: x[0],
            reverse=False
        )
        pass

    def getTimeline(self):
        self.f1()
        self.f2()
        self.f3()
        self.f4()
        self.f5()
        self.f6()
        self.f7()
        tl = []
        for t in self.timeline:
            if t[0] is not None:
                tl.append(t)
        self.timeline = tl
        pass

    def to_excel(self, path):
        wb = load_workbook(project_dir + '\\xxx公司发展历程.xlsx')
        sh = wb['公司历程']
        sh['A1'] = self.name
        for i in range(len(self.timeline)):
            sh.cell(i + 3, 1, self.timeline[i][0])
            sh.cell(i + 3, 2, self.timeline[i][1])
            sh.cell(i + 3, 3, self.timeline[i][2])
        wb.save(path)

    def f1(self):
        meta = '基本信息'
        d = self.bm.query_one(sql={'name': self.name, 'metaModel': meta})
        if d is None:
            return
        d = d['content']
        __ = d['工商信息']['#1']
        self.timeline.append([
            __['成立日期'], meta,
            '公司注册成立，注册资本{}{}'.format(
                __['注册资本']['金额'], __['注册资本']['单位']
            )])
        __ = d['变更记录']
        for b in __.values():
            self.timeline.append([
                b['变更日期'], meta,
                "公司发生{}：\n变更前:{}\n变更后:{}".format(
                    b['变更项目'],
                    b['变更前']['内容'],
                    b['变更后']['内容'],
                )
            ])
        __ = d['股东信息']
        for _ in __.values():
            if '认缴出资日期' in _.keys():
                self.timeline.append([
                    _['认缴出资日期'], meta,
                    '股东{}认缴出资{}{}'.format(
                        _['股东']['名称'],
                        _['认缴出资额']['金额'], _['认缴出资额']['单位']
                    )
                ])
            if '实缴出资日期' in _.keys():
                self.timeline.append([
                    _['实缴出资日期'], meta,
                    '股东{}实缴出资{}{}'.format(
                        _['股东']['名称'],
                        _['实缴出资额']['金额'], _['实缴出资额']['单位']
                    )
                ])
        __ = d['对外投资']
        for _ in __.values():
            if '融资日期' in _.keys():
                self.timeline.append([
                    _['融资日期'], meta,
                    '投资{}{}{}，所占比例{}'.format(
                        _['被投资企业']['名称'],
                        _['投资数额']['金额'], _['投资数额']['单位'],
                        _['投资比例']
                    )
                ])
        __ = d['建筑资质资格']
        for _ in __.values():
            self.timeline.append([
                _['发证日期'], meta,
                '获得建筑资质资格，证书名称{}'.format(
                    _['资质名称']
                )
            ])
        __ = d['股权变更']
        for _ in __.values():
            self.timeline.append([
                _['公示日期'], meta,
                '股东，股权比例由{}变更为{}'.format(
                    _['变更前股权比例'], _['变更后股权比例']
                )
            ])
        pass

    def f2(self):
        meta = '经营状况'
        d = self.bm.query_one(sql={'name': self.name, 'metaModel': meta})
        if d is None:
            return
        d = d['content']
        __ = d['产权交易']
        for _ in __.values():
            self.timeline.append([
                _['交易日期'], meta,
                '以{}{}转让{}给{}'.format(
                    _['转让价格']['金额'], _['转让价格']['单位'],
                    _['标的企业']['名称'], _['转让方']['名称']
                )
            ])
        __ = d['抽查检查']
        for _ in __.values():
            self.timeline.append([
                _['日期'], meta,
                '{}对公司进行{}检查，检查结果为{}'.format(
                    _['实施机关'], _['类型'],
                    _['结果']
                )
            ])
        __ = d['购地信息']
        for _ in __.values():
            self.timeline.append([
                _['合同签订日期'], meta,
                '购入位于{}，共{}{}土地，土地用途为{}，供地方式为{}'.format(
                    _['项目位置']['位置'], _['面积']['数量'], _['面积']['单位'],
                    _['土地用途'], _['供地方式']
                )
            ])
        __ = d['行政许可']
        for v in __.values():
            if '工商局' in v.keys():
                _ = v['工商局']
                self.timeline.append([
                    _['有效期自'], meta,
                    '由{}颁布行政许可"{}"'.format(
                        _['许可机关'], _['许可内容']
                    )
                ])
            if '信用中国' in v.keys():
                _ = v['信用中国']
                self.timeline.append([
                        _['决定日期'], meta,
                        '由{}颁布行政许可"{}"'.format(
                            _['许可机关'], _['决定文书号']
                        )
                    ])
        __ = d['进出口信用']
        for _ in __.values():
            self.timeline.append([
                _['注册日期'], meta,
                '在{}注册{}进出口信息'.format(
                    _['注册海关'], _['经营类别']
                )
            ])
        __ = d['双随机抽查']
        for _ in __.values():
            self.timeline.append([
                _['完成日期'], meta,
                '{}完成对公司的“{}”'.format(
                    _['抽查机关'], _['任务名称']
                )
            ])
        __ = d['招聘']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '发布招聘职位：{}，薪资：{}'.format(
                    _['职位']['职位'], _['月薪']
                )
            ])
        __ = d['招投标信息']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '{}：{}'.format(
                    _['项目分类'], _['描述']['描述']
                )
            ])
        __ = d['信用评级']
        for _ in __.values():
            self.timeline.append([
                _['评级日期'], meta,
                '公司被{}信用评级为{}'.format(
                    _['评级公司']['名称'], _['主体评级']
                )
            ])
        pass

    def f3(self):
        meta = '经营风险'
        d = self.bm.query_one(sql={'name': self.name, 'metaModel': meta})
        if d is None:
            return
        d = d['content']
        __ = d['动产抵押']
        for _ in __.values():
            self.timeline.append([
                _['登记日期'], meta,
                '涉及动产抵押，抵押权人：{}，债务人：{}，所有权或使用权归属{}，涉及金额{}{}'.format(
                    _['抵押权人']['名称'], _['债务人']['名称'],
                    _['所有权或使用权归属']['名称'],
                    _['被担保主债权数额']['金额'],
                    _['被担保主债权数额']['单位']
                )
            ])
        __ = d['公示催告']
        for _ in __.values():
            self.timeline.append([
                _['公告日期'], meta,
                '{}申请{}票据承兑，票面金额{}{}，持票人：{}'.format(
                    _['申请人']['名称'], _['票据类型'],
                    _['票面金额']['金额'],
                    _['票面金额']['单位'],
                    _['持票人']['名称'],
                )
            ])
        __ = d['股权出质']
        for _ in __.values():
            self.timeline.append([
                _['登记日期'], meta,
                '{}将{}出质给{}，涉及金额{}{}'.format(
                    _['出质人']['名称'],
                    _['标的企业']['名称'],
                    _['质权人']['名称'],
                    _['出质数额']['金额'],
                    _['出质数额']['单位'],
                )
            ])
        __ = d['行政处罚']['工商局']
        for _ in __.values():
            self.timeline.append([
                _['公示日期'] if _['公示日期'] is not None else _['决定日期'], meta,
                '公司因{}，{}对公司实施{}'.format(
                    _['违法行为类型'],
                    _['决定机关'],
                    _['处罚内容'],
                )
            ])
        __ = d['行政处罚']['税务局']
        for _ in __.values():
            self.timeline.append([
                _['处罚决定日期'], meta,
                '公司因{}，税务局对公司实施行政处罚'.format(
                    _['处罚事由'],
                )
            ])
        __ = d['行政处罚']['其他']
        for _ in __.values():
            self.timeline.append([
                _['处罚日期'], meta,
                '公司因{}，{}对公司实施行政处罚'.format(
                    _['处罚事由'],
                    _['处罚单位'],
                )
            ])
        __ = d['行政处罚']['信用中国']
        for _ in __.values():
            self.timeline.append([
                _['处罚日期'], meta,
                '公司因{}，{}对公司实施行政处罚'.format(
                    _['处罚事由'],
                    _['处罚机关'],
                )
            ])
        __ = d['环保处罚']
        for _ in __.values():
            self.timeline.append([
                _['处罚日期'], meta,
                '公司因{}，{}对公司实施行政处罚'.format(
                    _['违法类型'],
                    _['处罚单位'],
                )
            ])
        __ = d['简易注销']
        for _ in __.values():
            self.timeline.append([
                _['公告申请日期'], meta,
                '申请简易注销'
            ])
        __ = d['经营异常']
        for _ in __.values():
            self.timeline.append([
                _['列入日期'], meta,
                '因{}，被列入经营异常名单'.format(_['列入原因'])
            ])
            if '移出日期' in _.keys():
                self.timeline.append([
                    _['移出日期'], meta,
                    '因{}，被移出经营异常名单'.format(_['移出原因'])
                ])
        __ = d['破产重组']
        for _ in __.values():
            self.timeline.append([
                _['公开日期'], meta,
                '{}申请对{}进行破产重组'.format(
                    _['申请人']['名称'],
                    _['被申请人']['名称']
                )
            ])
        __ = d['欠税公告']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '公司涉及{}{}的欠税'.format(
                    _['欠税余额']['金额'],
                    _['欠税余额']['单位']
                )
            ])
        __ = d['税收违法']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '公司涉税收违法，案件性质：{}'.format(
                    _['案件性质'],
                )
            ])
        __ = d['司法拍卖']
        for _ in __.values():
            self.timeline.append([
                _['拍卖时间'], meta,
                '{}'.format(
                    _['标题'],
                )
            ])
        __ = d['土地抵押']
        for _ in __.values():
            self.timeline.append([
                _['抵押起止日期'].split('至')[0].strip().replace('\n', ''), meta,
                '{}将{}{}土地抵押给{}，抵押金额{}{}，土地坐落于{}'.format(
                    _['抵押人']['名称'],
                    _['抵押面积']['数额'], _['抵押面积']['单位'],
                    _['抵押权人']['名称'],
                    _['抵押金额']['金额'], _['抵押金额']['单位'],
                    _['位置']
                )
            ])
        __ = d['询价评估']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '公司就{}发布询价评估，询价结果{}{}'.format(
                    _['标的物']['名称'],
                    _['询价结果']['金额'], _['询价结果']['单位'],
                )
            ])
        __ = d['严重违法']
        for _ in __.values():
            self.timeline.append([
                _['列入日期'], meta,
                '因{}，被列入严重违法名单'.format(_['列入原因'])
            ])
            if '移出日期' in _.keys():
                self.timeline.append([
                    _['移出日期'], meta,
                    '因{}，被移出严重违法名单'.format(_['移出原因'])
                ])
        __ = d['注销备案']
        if '清算组备案信息' in __.keys():
            _ = __['清算组备案信息']
            self.timeline.append([
                _['清算组备案日期'], meta,
                '公司因{}成立清算组'.format(
                    _['注销原因'],
                )
            ])
        pass

    def f4(self):
        meta = '企业发展'
        d = self.bm.query_one(sql={'name': self.name, 'metaModel': meta})
        if d is None:
            return
        d = d['content']
        __ = d['企业业务']
        for _ in __.values():
            self.timeline.append([
                _['成立日期'], meta,
                '公司开发或生产{}产品'.format(
                    _['产品名']['名称'],
                )
            ])
        __ = d['融资信息']
        for _ in __.values():
            self.timeline.append([
                _['日期'], meta,
                '公司向{}融资{}{}'.format(
                    _['投资方']['名称'],
                    _['金额']['金额'], _['金额']['单位'],
                )
            ])
        pass

    def f5(self):
        meta = '知识产权'
        d = self.bm.query_one(sql={'name': self.name, 'metaModel': meta})
        if d is None:
            return
        d = d['content']
        __ = d['软件著作权']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'] if _['发布日期'] is not None else _['登记批准日期'], meta,
                '公司申请了{}的软件著作权'.format(
                    _['软件名称'],
                )
            ])
        __ = d['商标信息']
        for _ in __.values():
            self.timeline.append([
                _['申请日期'], meta,
                '公司申请了商标：{}'.format(
                    _['商标']['名称'],
                )
            ])
        __ = d['网站信息']
        for _ in __.values():
            self.timeline.append([
                _['审核日期'], meta,
                '公司申请了网站备案：{}'.format(
                    _['名称'],
                )
            ])
        __ = d['证书信息']
        for _ in __.values():
            self.timeline.append([
                _['发证日期'], meta,
                '公司获得了{}证书'.format(
                    _['证书']['名称'],
                )
            ])
        __ = d['专利信息']
        for _ in __.values():
            self.timeline.append([
                _['公开日期'], meta,
                '公司申请了{}专利'.format(
                    _['专利']['名称'],
                )
            ])
        __ = d['作品著作权']
        for _ in __.values():
            self.timeline.append([
                _['首次发表日期'], meta,
                '公司申请了{}的作品著作权'.format(
                    _['作品名称'],
                )
            ])
        pass

    def f6(self):
        meta = '法律诉讼'
        d = self.bm.query_one(sql={'name': self.name, 'metaModel': meta})
        if d is None:
            return
        d = d['content']
        try:
            __ = d['被执行人']
            for _ in __.values():
                self.timeline.append([
                    _['立案日期'], meta,
                    '公司被列为“被执行人”，执行金额{}{}'.format(
                        _['执行标的']['金额'], _['执行标的']['单位'],
                    )
                ])
        except Exception as e:
            ExceptionInfo(e)
        __ = d['裁判文书']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '公司涉及的“{}”案件：{}，裁决结果：{}'.format(
                    _['案由'], _['案号'], _['裁判文书']['标题'],
                )
            ])
        __ = d['法院公告']
        for _ in __.values():
            self.timeline.append([
                _['刊登日期'], meta,
                '法院公告：{}与{}的“{}”'.format(
                    _['被告人/被告/被上诉人/被申请人']['名称'],
                    _['公诉人/原告/上诉人/申请人']['名称'],
                    _['案由']
                )
            ])
        # __ = d['股权冻结']
        # for _ in __.values():
        #     self.timeline.append([
        #         _['刊登日期'],
        #         '{}的{}被执行股权冻结，冻结股权数额{}{}'.format(
        #             _['被执行人']['名称'],
        #             _['标的企业']['名称'],
        #             _['股权数额']['金额'], _['股权数额']['单位']
        #         )
        #     ])
        __ = d['开庭公告']
        for _ in __.values():
            self.timeline.append([
                _['开庭时间'], meta,
                '开庭公告：{}与{}的“{}”'.format(
                    _['被告人/被告/被上诉人/被申请人']['名称'],
                    _['公诉人/原告/上诉人/申请人']['名称'],
                    _['案由']
                )
            ])
        __ = d['立案信息']
        for _ in __.values():
            self.timeline.append([
                _['立案日期'], meta,
                '立案信息：{}与{}的“{}”'.format(
                    _['被告人/被告/被上诉人/被申请人']['名称'],
                    _['公诉人/原告/上诉人/申请人']['名称'],
                    _['案由']
                )
            ])
        __ = d['失信被执行人']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '公司被列为失信被执行人，履行情况：{}'.format(
                    _['履行情况']
                )
            ])
        __ = d['送达公告']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '送达公告：{}与{}的“{}”'.format(
                    _['被告人/被告/被上诉人/被申请人']['名称'],
                    _['公诉人/原告/上诉人/申请人']['名称'],
                    _['案由']
                )
            ])
        __ = d['限制高消费']
        for _ in __.values():
            self.timeline.append([
                _['发布日期'], meta,
                '因“{}”，{}申请对{}实施限制高消费'.format(
                    _['案号']['名称'],
                    _['申请人']['名称'],
                    _['限消令对象']['名称'],
                )
            ])
        __ = d['终本案件']
        for _ in __.values():
            self.timeline.append([
                _['终本日期'], meta,
                '终本案件：执行标的{}{}，未履行{}{}'.format(
                    _['执行标的']['金额'], _['执行标的']['单位'],
                    _['未履行金额']['金额'], _['未履行金额']['单位'],
                )
            ])
        pass

    def f7(self):
        meta = '公司新闻'
        d = self.bm.query_one(sql={'name': self.name, 'metaModel': meta})
        if d is None:
            return
        d = d['content']
        __ = d['企业公告']
        for _ in __.values():
            self.timeline.append([
                _['日期'], meta,
                '发布企业公告：{}'.format(
                    _['标题'],
                )
            ])
        __ = d['相关公告']
        for _ in __.values():
            self.timeline.append([
                _['日期'], meta,
                '相关公告：{}'.format(
                    _['标题'],
                )
            ])
        __ = d['新闻舆情']
        for _ in __.values():
            self.timeline.append([
                _['发布时间'], meta,
                '公司新闻：{}'.format(
                    _['新闻标题'],
                )
            ])
        pass


n = '重庆一胜特工模材料有限公司'
Timeline(n).to_excel('D:\graph_data\【发展历程】{}.xlsx'.format(n))